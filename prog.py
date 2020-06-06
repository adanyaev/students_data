import pandas as pd
import datetime
import xlrd
import xlwt
import openpyxl
#D:\19ПИ-3.xlsx - это путь к файлу
class STUDENT(object):
    def __init__(self, name='', group='', results=[]):
        self.name = name
        self.group = group
        self.results = results

    def show_student(self):
        discription = (str(self.name) + " " + str(self.group) + " middle point is " + str(
            self.midpoint) + " result point is " + str(self.respoint))
        print(discription)
def Current_student_grades(*file_names):
    students = []
    for file_num in range(int(len(file_names))):
        file_name = file_names[file_num]
        rb = xlrd.open_workbook(file_name, formatting_info=False)
        if rb == 0:
            print(str("Error of opening: " + str(file_name)))
            continue
        sheet = rb.sheet_by_index(0)            #выбираем активный лист
        row_number = sheet.nrows
        col_number = sheet.ncols
        group = sheet.cell(rowx=1, colx=sheet.ncols).value
        subject = sheet.cell(rowx=2, colx=sheet.ncols).value
        i = 2
        kolvo = 0
        while sheet.cell(rowx=0, colx=i).value != 'Средний балл' and sheet.cell(rowx=1, colx=i).value != 'Middle point':
            kolvo += 1
            i += 1
        k = 0
        if row_number > 0:
            for i in range(1, row_number):
                array = [0]*kolvo
                j = 1
                k = 0
                flag = 0
                dict_ = {}
                while sheet.cell(rowx=0, colx=j + 1).value != 'Средний балл' and sheet.cell(rowx=1,
                                                                                            colx=j).value != 'Middle point':
                    cell = sheet.cell(rowx=0, colx=j + 1)
                    j += 1
                    if cell.ctype == 3:
                        year, month, day, hour, minute, second = xlrd.xldate_as_tuple(cell.value, rb.datemode)
                        value = datetime.datetime(year, day, month)
                    else:
                        value = cell.value
                    mark = str(sheet.cell(rowx=i, colx=j).value)
                    dict_.update({str(value): mark})
                dict_.update({sheet.cell(rowx=0, colx=kolvo + 2).value: sheet.cell(rowx=i, colx=kolvo + 2).value})
                dict_.update({sheet.cell(rowx=0, colx=kolvo + 3).value: sheet.cell(rowx=i, colx=kolvo + 3).value})
                for m in students:
                    if students[m].name == sheet.cell(rowx=i, colx=1).value:
                        flag = 1
                        students[m].results.append(subject)
                        students[m].results.append(dict_)
                if flag == 0:
                    results = []
                    results.append(subject)
                    results.append(dict_)
                    Student = STUDENT(sheet.cell(rowx=i, colx=1).value, group, results)
                    students.append(Student)
                    results.clear()
        else:
            print('Empty file' + file_name)
            continue
        return students
#print('Enter the way to your file(with name of this file): ')
#way_ = str(input())
#arr = []
#arr = Current_student_grades(way_)
#i = 0
#for i in range (len(arr)):
 #+   arr[i].show_student()